from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from numbers import Number
import re
from typing import TYPE_CHECKING

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet
    import pandas as pd

# Default Excel table style name used for all exported tables.
DEFAULT_EXCEL_TABLE_STYLE = "TableStyleMedium9"

# Default numeric format for non-coordinate numeric values.
DEFAULT_NUMERIC_FORMAT = "0.000"

# Default integer format for unique identifiers.
DEFAULT_INTEGER_FORMAT = "0"

# Higher precision format for latitude and longitude values.
DEFAULT_COORDINATE_FORMAT = "0.000000"

# Date-only display format for datetime values in exported sheets.
DEFAULT_DATE_FORMAT = "yyyy-mm-dd"

# Header cell fill color used for table headers.
DEFAULT_HEADER_FILL = "FFD9E1F2"

# Header font should be bold for readability.
DEFAULT_HEADER_FONT_BOLD = True

# Header alignment used for table header cells.
DEFAULT_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# Default alignment for data cells, including wrap text for long values.
DEFAULT_CELL_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

# Minimum and maximum column widths used when auto-sizing.
DEFAULT_COLUMN_WIDTH_MIN = 12.0
DEFAULT_COLUMN_WIDTH_MAX = 40.0


@dataclass(frozen=True)
class ExcelTableOptions:
    """Options for creating a styled Excel table.

    Attributes:
        table_name: Optional table display name.
        table_style: Excel table style name.
        min_column_width: Minimum width in Excel units.
        max_column_width: Maximum width in Excel units.
        integer_columns: Column names whose numeric values should use integer formatting.
        coordinate_columns: Column names whose numeric values should use coordinate precision.
    """

    table_name: str | None = None
    table_style: str = DEFAULT_EXCEL_TABLE_STYLE
    min_column_width: float = DEFAULT_COLUMN_WIDTH_MIN
    max_column_width: float = DEFAULT_COLUMN_WIDTH_MAX
    integer_columns: frozenset[str] | None = None
    coordinate_columns: frozenset[str] | None = None


def _sanitize_table_name(name: str) -> str:
    """Normalize a raw table name so it is valid in Excel.

    This replaces invalid characters with underscores and ensures the name
    does not begin with a digit, which is required by Excel table naming rules.

    Args:
        name: Raw worksheet title.

    Returns:
        A valid Excel table name.
    """
    sanitized = re.sub(r"[^A-Za-z0-9_]", "_", name)
    if not sanitized:
        sanitized = "Table1"
    if sanitized[0].isdigit():
        sanitized = f"T_{sanitized}"
    return sanitized


def _get_column_widths(
    dataframe: pd.DataFrame,
    min_width: float,
    max_width: float,
) -> dict[int, float]:
    """Compute column widths from dataframe contents and clamp them to bounds.

    Each width is based on the longest text in the column, then bounded by
    the provided minimum and maximum widths.

    Args:
        dataframe: Dataframe to measure.
        min_width: Minimum width in Excel units.
        max_width: Maximum width in Excel units.

    Returns:
        Mapping of 1-based column indexes to computed widths.
    """
    widths: dict[int, float] = {}
    for index, column in enumerate(dataframe.columns, start=1):
        header_width = len(str(column))
        max_text_width = header_width
        for value in dataframe[column].astype(str):
            max_text_width = max(max_text_width, len(value))

        width = max(min_width, min(max_text_width + 2.0, max_width))
        widths[index] = width
    return widths


def _style_header_row(worksheet: Worksheet) -> None:
    """Apply header styling to the first row (header) of the worksheet.

    Args:
        worksheet: Worksheet instance to style.
    """
    for cell in worksheet[1]:
        cell.font = Font(bold=DEFAULT_HEADER_FONT_BOLD)
        cell.fill = PatternFill(fill_type="solid", fgColor=DEFAULT_HEADER_FILL)
        cell.alignment = DEFAULT_HEADER_ALIGNMENT


def _style_data_cells(
    worksheet: Worksheet,
    dataframe: pd.DataFrame,
    options: ExcelTableOptions,
) -> None:
    """Apply cell formatting to numeric, date, and text cells.

    Args:
        worksheet: Worksheet instance to format.
        dataframe: Dataframe used to define the table range.
        options: Styling options that include column-specific formats.
    """
    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
        max_col=dataframe.shape[1],
    ):
        for column_index, cell in enumerate(row, start=1):
            cell.alignment = DEFAULT_CELL_ALIGNMENT
            if isinstance(cell.value, (datetime, date)) and not isinstance(
                cell.value,
                bool,
            ):
                cell.number_format = DEFAULT_DATE_FORMAT
            elif isinstance(cell.value, Number) and not isinstance(cell.value, bool):
                column_name = dataframe.columns[column_index - 1]
                if (
                    options.integer_columns is not None
                    and column_name in options.integer_columns
                ):
                    cell.number_format = DEFAULT_INTEGER_FORMAT
                elif (
                    options.coordinate_columns is not None
                    and column_name in options.coordinate_columns
                ):
                    cell.number_format = DEFAULT_COORDINATE_FORMAT
                else:
                    cell.number_format = DEFAULT_NUMERIC_FORMAT


def _apply_column_widths(
    worksheet: Worksheet,
    dataframe: pd.DataFrame,
    min_width: float,
    max_width: float,
) -> None:
    """Set column widths on the worksheet based on data content.

    Args:
        worksheet: Worksheet instance to adjust.
        dataframe: Dataframe used to determine widths.
        min_width: Minimum width in Excel units.
        max_width: Maximum width in Excel units.
    """
    widths = _get_column_widths(dataframe, min_width, max_width)
    for index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(index)].width = width


def create_excel_table(
    worksheet: Worksheet,
    dataframe: pd.DataFrame,
    options: ExcelTableOptions | None = None,
) -> None:
    """Create and style an Excel table from a dataframe.

    Args:
        worksheet: Worksheet instance to populate.
        dataframe: Dataframe to write as a table.
        options: Optional table styling options.
    """
    if options is None:
        options = ExcelTableOptions()

    table_name = options.table_name or f"Table_{worksheet.title}"
    table_name = _sanitize_table_name(table_name)
    end_column = get_column_letter(dataframe.shape[1])
    end_row = dataframe.shape[0] + 1
    table_ref = f"A1:{end_column}{end_row}"

    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name=options.table_style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    worksheet.add_table(table)
    _style_header_row(worksheet)
    _style_data_cells(worksheet, dataframe, options)
    _apply_column_widths(
        worksheet,
        dataframe,
        options.min_column_width,
        options.max_column_width,
    )
