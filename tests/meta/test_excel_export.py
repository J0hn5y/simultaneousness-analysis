from pathlib import Path

from openpyxl import load_workbook
import pandas as pd

from simultaneousness_analysis.meta.table import MetaTable


def _create_meta_table(tmp_path: Path) -> tuple[MetaTable, Path]:
    """Build and export a sample meta table for worksheet validation."""
    meta_table = MetaTable.__new__(MetaTable)
    meta_table.data_path = tmp_path
    meta_table.table = pd.DataFrame(
        {
            "stations_id": [4466, 2907],
            "station_name": ["Test Station", "Another Station"],
            "measurement": [1.23456, 7.0],
            "latitude": [54.5275, 54.7903],
            "longitude": [9.5487, 8.9514],
            "from_date": ["2000-01-01", "2001-05-13"],
            "to_date": ["2000-12-31", "2001-12-31"],
            "notes": [
                "Short text",
                "A longer piece of text that should wrap automatically.",
            ],
        },
    )
    meta_table.table["from_date"] = pd.to_datetime(meta_table.table["from_date"])
    meta_table.table["to_date"] = pd.to_datetime(meta_table.table["to_date"])
    meta_table.table.index = pd.Index(["path-a", "path-b"], name="path")

    output_path = tmp_path / "meta_table.xlsx"
    meta_table.export(suffix="xlsx", path=output_path)
    return meta_table, output_path


def _load_meta_table_worksheet(output_path: Path):
    """Load the exported worksheet for assertions."""
    workbook = load_workbook(output_path)
    return workbook["Meta Table"]


def test_meta_table_xlsx_export_creates_excel_table(tmp_path: Path) -> None:
    """Verify an Excel table is created and named correctly."""
    _, output_path = _create_meta_table(tmp_path)
    worksheet = _load_meta_table_worksheet(output_path)

    assert len(worksheet._tables) == 1
    table = next(iter(worksheet._tables.values()))
    assert table.ref == "A1:I3"
    assert table.displayName.startswith("Table")


def test_meta_table_xlsx_export_applies_header_styles(tmp_path: Path) -> None:
    """Verify header formatting is applied to the exported sheet."""
    _, output_path = _create_meta_table(tmp_path)
    worksheet = _load_meta_table_worksheet(output_path)

    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, max_col=9))
    for cell in header_row:
        assert cell.font.bold is True
        assert cell.fill.fill_type == "solid"


def test_meta_table_xlsx_export_formats_integer_columns(tmp_path: Path) -> None:
    """Verify integer columns use integer formatting."""
    _, output_path = _create_meta_table(tmp_path)
    worksheet = _load_meta_table_worksheet(output_path)

    assert worksheet["B2"].number_format == "0"
    assert worksheet["B3"].number_format == "0"


def test_meta_table_xlsx_export_formats_default_numeric_columns(tmp_path: Path) -> None:
    """Verify default numeric columns use standard decimal formatting."""
    _, output_path = _create_meta_table(tmp_path)
    worksheet = _load_meta_table_worksheet(output_path)

    assert worksheet["D2"].number_format == "0.000"
    assert worksheet["D3"].number_format == "0.000"


def test_meta_table_xlsx_export_formats_coordinate_columns(tmp_path: Path) -> None:
    """Verify coordinate columns use higher precision formatting."""
    _, output_path = _create_meta_table(tmp_path)
    worksheet = _load_meta_table_worksheet(output_path)

    assert worksheet["E2"].number_format == "0.000000"
    assert worksheet["E3"].number_format == "0.000000"
    assert worksheet["F2"].number_format == "0.000000"
    assert worksheet["F3"].number_format == "0.000000"


def test_meta_table_xlsx_export_formats_date_columns(tmp_path: Path) -> None:
    """Verify date columns are formatted using the date format."""
    _, output_path = _create_meta_table(tmp_path)
    worksheet = _load_meta_table_worksheet(output_path)

    assert worksheet["G2"].number_format == "yyyy-mm-dd"
    assert worksheet["G3"].number_format == "yyyy-mm-dd"
    assert worksheet["H2"].number_format == "yyyy-mm-dd"
    assert worksheet["H3"].number_format == "yyyy-mm-dd"


def test_meta_table_xlsx_export_wraps_text_in_notes(tmp_path: Path) -> None:
    """Verify text wrapping is enabled for the notes column."""
    _, output_path = _create_meta_table(tmp_path)
    worksheet = _load_meta_table_worksheet(output_path)

    assert worksheet["I2"].alignment.wrap_text is True
    assert worksheet["I3"].alignment.wrap_text is True


def test_meta_table_xlsx_export_sets_column_widths_within_bounds(
    tmp_path: Path,
) -> None:
    """Verify all column widths stay inside the configured bounds."""
    _, output_path = _create_meta_table(tmp_path)
    worksheet = _load_meta_table_worksheet(output_path)

    column_widths = [
        worksheet.column_dimensions[col].width
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
    ]

    assert all(width is not None for width in column_widths)
    assert all(12.0 <= width <= 40.0 for width in column_widths)
